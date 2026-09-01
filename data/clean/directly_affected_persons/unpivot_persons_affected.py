"""
Converts the wide-format "directly affected persons" Excel file
(Country | 2005 | 2006 | ... | 2023) into long format
(Pacific Island Countries and Territories | Year | Value),
with an AutoFilter dropdown ready on every column.

Run from inside scripts/ — expects:
  ../source/directly_affected_persons.xlsx
Writes:
  ../output/directly_affected_persons_LONG.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SOURCE_PATH = '../source/directly_affected_persons.xlsx'
OUTPUT_PATH = '../output/directly_affected_persons_LONG.xlsx'
SHEET_NAME = 'Sheet1'
VALUE_COLUMN_NAME = 'Persons Affected'
COUNTRY_COLUMN_NAME = 'Pacific Island Countries and Territories'

# Region header rows in the source sheet that aren't actual countries
REGION_HEADERS = {'Melanesia', 'Micronesia', 'Polynesia'}


def unpivot(source_path, sheet, value_name, out_path):
    df = pd.read_excel(source_path, sheet_name=sheet)
    year_cols = [c for c in df.columns if c != 'Country']

    # Drop the region-header rows (Melanesia / Micronesia / Polynesia) —
    # they carry no data of their own, just group the countries beneath them
    df = df[~df['Country'].isin(REGION_HEADERS)].copy()

    long_df = df.melt(
        id_vars=['Country'],
        value_vars=year_cols,
        var_name='Year',
        value_name=value_name,
    )
    long_df['Year'] = long_df['Year'].astype(int)
    long_df = long_df.rename(columns={'Country': COUNTRY_COLUMN_NAME})
    long_df = long_df.sort_values([COUNTRY_COLUMN_NAME, 'Year']).reset_index(drop=True)
    long_df.to_excel(out_path, sheet_name='Long Format', index=False)
    return long_df


def style_file(path):
    wb = openpyxl.load_workbook(path)
    ws = wb['Long Format']

    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial')

    widths = {'A': 34, 'B': 10, 'C': 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 30

    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
    ws.freeze_panes = "A2"

    wb.save(path)


if __name__ == '__main__':
    df = unpivot(SOURCE_PATH, SHEET_NAME, VALUE_COLUMN_NAME, OUTPUT_PATH)
    style_file(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH} — {len(df)} rows")
